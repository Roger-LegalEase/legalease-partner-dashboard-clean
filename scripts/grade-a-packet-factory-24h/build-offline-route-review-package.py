#!/usr/bin/env python3
"""The offline route review package: PDFs, a prefilled workbook, a manifest, one ZIP.

WHY OFFLINE. A reviewer reading a legal packet needs to zoom, scroll back, sit
with a page and put it down. A hosted instrument is worse at all four, and it
binds the reviewer to a session. This produces the actual PDFs the participant
would receive, a workbook with one row per page, and a manifest that ties every
file and every page to a digest — so a completed review can be checked against
the bytes it was made of, long after the session is gone.

WHAT MAKES A ROW TRUSTWORTHY. Every row carries the runtime route id, the
artifact's SHA-256 and that page's own rendered SHA-256, all recomputed here
from the files being shipped rather than copied from a record. A review that
comes back names bytes: if an artifact is rebuilt and its digest moves, the rows
that named the old digest visibly no longer apply, instead of being silently
inherited by different bytes.

WHAT IT REFUSES. A route whose artifact is missing, whose digest disagrees with
the record, or which the caller did not name is not shipped. A partial package
that looks complete is the failure mode this exists to avoid.

    python3 scripts/grade-a-packet-factory-24h/build-offline-route-review-package.py \
        --family rcap-ks-custom-pleading --state Kansas \
        --routes ks-12-4516-municipal,ks-12-4516a-municipal-arrest \
        --out /tmp/review-packages
"""
from __future__ import annotations

import argparse
import hashlib
import json
import shutil
import subprocess
import zipfile
from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

ROOT = Path(__file__).resolve().parents[2]
HEAD_FILL = PatternFill("solid", fgColor="2F5D50")
HEAD_FONT = Font(color="FFFFFF", bold=True, size=10)
WRAP = Alignment(vertical="top", wrap_text=True)
TOP = Alignment(vertical="top")

ISSUE_CATEGORIES = [
    "", "Wrong or missing content", "Clipped or overlapping text", "Box or field wrongly marked",
    "Blank that should be filled", "Filled that should be blank", "Wrong court or recipient",
    "Wrong fee or waiver statement", "Missing instruction", "Formatting only", "Other",
]


def sha256(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as fh:
        for chunk in iter(lambda: fh.read(1 << 20), b""):
            h.update(chunk)
    return h.hexdigest()


def render_pages(pdf: Path, dest: Path, dpi: int) -> list[Path]:
    dest.mkdir(parents=True, exist_ok=True)
    subprocess.run(["pdftoppm", "-png", "-r", str(dpi), str(pdf), str(dest / "page")], check=True)
    return sorted(dest.glob("page-*.png"))


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--family", required=True)
    ap.add_argument("--state", required=True)
    ap.add_argument("--routes", required=True, help="comma-separated route ids as they appear in routeArtifacts")
    ap.add_argument("--out", default="/tmp/review-packages")
    ap.add_argument("--dpi", type=int, default=150)
    args = ap.parse_args()

    master = json.loads((ROOT / "data/rcap-grade-a/packet-factory-24h/MASTER_QUEUE.json").read_text())
    census_doc = json.loads((ROOT / "data/rcap-grade-a/route-obligation-census-candidate/route-obligation-candidate.json").read_text())
    census = next(v for v in census_doc.values() if isinstance(v, list) and len(v) > 50)

    fam = next((f for f in master["families"] if f["familyId"] == args.family), None)
    if fam is None:
        raise SystemExit(f"REFUSED: {args.family} is not in the master queue")
    art = json.loads((ROOT / fam["directory"] / "reports/rendered-artifacts.json").read_text())
    route_artifacts = art.get("routeArtifacts") or []
    if not route_artifacts:
        raise SystemExit(f"REFUSED: {args.family} declares no routeArtifacts; the per-route repair has not landed")

    # The runtime route id, from the crosswalk rather than composed here.
    runtime_of: dict[str, str] = {}
    for row in census:
        pathway = row.get("runtimePathwayId")
        track = row.get("trackId")
        if pathway and track:
            runtime_of[track] = f"{row.get('jurisdiction')}:{pathway}"

    routes = [r.strip() for r in args.routes.split(",") if r.strip()]
    pkg = Path(args.out) / f"{args.state}_route_review_{date.today().isoformat()}"
    if pkg.exists():
        shutil.rmtree(pkg)
    (pkg / "packets").mkdir(parents=True)
    work = pkg / ".pages"

    rows, manifest_files = [], []
    for route in routes:
        runtime_id = runtime_of.get(route)
        if not runtime_id:
            raise SystemExit(f"REFUSED: {route} has no runtime route id in the crosswalk; it may not be shipped for review")
        for fixture in ("canonical", "boundary"):
            rec = next((r for r in route_artifacts
                        if (r.get("route") or r.get("routeKey")) == route and r.get("fixture") == fixture), None)
            if rec is None:
                raise SystemExit(f"REFUSED: {route}/{fixture} has no route artifact")
            src = ROOT / rec["file"]
            if not src.exists():
                raise SystemExit(f"REFUSED: {rec['file']} is absent from disk")
            digest = sha256(src)
            if rec.get("sha256") and rec["sha256"] != digest:
                raise SystemExit(f"REFUSED: {route}/{fixture} on disk is {digest[:12]} but the record declares {rec['sha256'][:12]}")

            filename = f"{args.state}_{route}_{fixture}.pdf"
            shutil.copy2(src, pkg / "packets" / filename)
            pages = render_pages(src, work / route / fixture, args.dpi)
            if rec.get("pageCount") and len(pages) != rec["pageCount"]:
                raise SystemExit(f"REFUSED: {route}/{fixture} rendered {len(pages)} pages, record declares {rec['pageCount']}")

            manifest_files.append({
                "runtimeRouteId": runtime_id, "route": route, "fixture": fixture,
                "filename": f"packets/{filename}", "sourcePath": rec["file"],
                "sha256": digest, "byteLength": src.stat().st_size, "pageCount": len(pages),
                "pageSha256": [sha256(p) for p in pages],
            })
            for i, page in enumerate(pages, start=1):
                rows.append([
                    runtime_id, filename, i, len(pages), digest, sha256(page),
                    "UNREVIEWED", "", "", "", "",
                ])

    # ---- the workbook ------------------------------------------------------
    wb = Workbook()
    ws = wb.active
    ws.title = f"{args.state} page review"
    headers = ["Canonical runtime route ID", "Artifact filename", "Page", "Total pages",
               "Artifact SHA-256", "Rendered-page SHA-256", "Result", "Issue category",
               "Reviewer comment", "Reviewer name", "Review date"]
    ws.append(headers)
    for i, w in enumerate([56, 46, 6, 11, 68, 68, 14, 26, 52, 20, 14], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    for cell in ws[1]:
        cell.fill, cell.font, cell.alignment = HEAD_FILL, HEAD_FONT, WRAP
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 30
    for r in rows:
        ws.append(r)
    for row in ws.iter_rows(min_row=2):
        for c in row:
            c.alignment = WRAP if c.column in (5, 6, 9) else TOP

    last = len(rows) + 1
    result_dv = DataValidation(type="list", formula1='"PASS,FAIL,COMMENT,UNREVIEWED"', allow_blank=False)
    issue_dv = DataValidation(type="list", formula1=f'"{",".join(c for c in ISSUE_CATEGORIES if c)}"', allow_blank=True)
    ws.add_data_validation(result_dv)
    ws.add_data_validation(issue_dv)
    result_dv.add(f"G2:G{last}")
    issue_dv.add(f"H2:H{last}")

    ws2 = wb.create_sheet("How to use this")
    for label, text in [
        ("What you are reviewing", f"Every page of the {args.state} packets a participant on these routes would actually receive. Nothing from any other route is in this package."),
        ("What to do", "Open the PDF named in each row, look at that page, and set Result to PASS or FAIL. Use COMMENT for something worth recording that is not a failure. Leave UNREVIEWED on any page you have not looked at — it will not be counted as a pass."),
        ("Why the hashes are here", "Each row names the artifact's SHA-256 and that page's own rendered SHA-256. Your review is bound to those exact bytes. If a packet is rebuilt and its digest changes, the rows naming the old digest no longer apply and will not be silently carried over."),
        ("One page fails, one route stops", "A failure removes that route from the cohort, not the others. Clean routes are not held behind it."),
        ("What this does not do", "A completed review satisfies one Grade A condition. It opens no route, sets no price and grants no payment or sponsorship eligibility."),
    ]:
        ws2.append([label, text])
    ws2.column_dimensions["A"].width = 34
    ws2.column_dimensions["B"].width = 110
    for row in ws2.iter_rows():
        row[0].font = Font(bold=True, size=10)
        for c in row:
            c.alignment = WRAP

    workbook_name = f"{args.state}_page_review.xlsx"
    wb.save(pkg / workbook_name)

    # ---- the manifest ------------------------------------------------------
    head = subprocess.run(["git", "rev-parse", "HEAD"], cwd=ROOT, capture_output=True, text=True).stdout.strip()
    manifest = {
        "schemaVersion": "rcap-offline-route-review-manifest/v1",
        "state": args.state, "familyId": args.family, "capturedAtCommit": head,
        "routes": routes, "runtimeRouteIds": [runtime_of[r] for r in routes],
        "totalPages": len(rows), "files": manifest_files,
        "workbook": {"filename": workbook_name},
        "whatAReturnedReviewProves": "That a named person looked at every page of the exact bytes named here and judged each fit to file or not. It proves nothing about any other bytes or any other route.",
        "whatItDoesNotGrant": "No commercial authority, no price, no payment or sponsorship eligibility. It satisfies one Grade A condition and none of the others.",
    }
    (pkg / "MANIFEST.json").write_text(json.dumps(manifest, indent=2) + "\n")
    (pkg / workbook_name).with_suffix(".xlsx")  # named for the manifest above
    manifest["workbook"]["sha256"] = sha256(pkg / workbook_name)
    (pkg / "MANIFEST.json").write_text(json.dumps(manifest, indent=2) + "\n")

    shutil.rmtree(work)
    zip_path = Path(args.out) / f"{pkg.name}.zip"
    with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as z:
        for p in sorted(pkg.rglob("*")):
            if p.is_file():
                z.write(p, p.relative_to(pkg.parent))

    print(f"wrote {zip_path}")
    print(f"  {len(manifest_files)} PDF(s) across {len(routes)} route(s), {len(rows)} page rows")
    for f in manifest_files:
        print(f"    {f['runtimeRouteId']}  {f['fixture']:9} {f['pageCount']:2}pp  {f['sha256'][:16]}…")
    print(f"  zip sha256 {sha256(zip_path)}")


if __name__ == "__main__":
    main()
