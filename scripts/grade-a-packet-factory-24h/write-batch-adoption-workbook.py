#!/usr/bin/env python3
"""Render the batch-adoption package as a workbook the decision owner can work in.

The JSON package is the record; this is the surface for the decision. It carries
one row per family with the evidence attached and an ADOPT / EXCLUDE / QUESTION
column, plus the exceptions and the eight open questions on their own sheets so
the exclusions are as visible as the list being adopted.

Nothing here decides anything. It is prepared for the owner and applied by
nobody until the owner returns it.

    python3 scripts/grade-a-packet-factory-24h/write-batch-adoption-workbook.py \
        --out /tmp/batch-adoption/LegalEase_Batch_Adoption_2026-09-02.xlsx
"""
from __future__ import annotations

import argparse
import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[2]
PACKAGE = ROOT / "data/rcap-grade-a/legal-decisions/BATCH_ADOPTION_PACKAGE_2026-09-02.json"

HEAD_FILL = PatternFill("solid", fgColor="2F5D50")
HEAD_FONT = Font(color="FFFFFF", bold=True, size=10)
WRAP = Alignment(vertical="top", wrap_text=True)
TOP = Alignment(vertical="top")


def flat(value):
    """Anything the record holds, rendered as a cell.

    Several fields in the delta record are sometimes a string and sometimes an
    object carrying the same text with an id beside it, because different
    classification paths wrote them. A workbook is not the place to discover
    that, so every value is coerced here and an object is rendered by its own
    question or text rather than as JSON a reader would have to parse.
    """
    if value is None:
        return ""
    if isinstance(value, (str, int, float, bool)):
        return value
    if isinstance(value, dict):
        for key in ("question", "text", "summary", "delta", "id"):
            if key in value and isinstance(value[key], str):
                extra = value.get("id") if key != "id" else None
                return f"{value[key]}" + (f"  [{extra}]" if extra and extra != value[key] else "")
        return json.dumps(value, ensure_ascii=False)
    if isinstance(value, list):
        return "\n".join(str(flat(v)) for v in value)
    return str(value)


def sheet(wb, title, headers, widths):
    ws = wb.create_sheet(title) if wb.sheetnames != ["Sheet"] else wb.active
    ws.title = title
    ws.append(headers)
    _append = ws.append
    ws.append = lambda row: _append([flat(v) for v in row])
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    for cell in ws[1]:
        cell.fill, cell.font, cell.alignment = HEAD_FILL, HEAD_FONT, WRAP
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 30
    return ws


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--out", required=True)
    args = ap.parse_args()
    pkg = json.loads(PACKAGE.read_text())

    wb = Workbook()

    # ---- the decision sheet -------------------------------------------------
    ws = sheet(
        wb, "Adopt (53 families)",
        ["Decision (ADOPT / EXCLUDE / QUESTION)", "Family", "State", "Routes",
         "Runtime-represented routes", "Delivery type", "Controlling legal-design record",
         "Shipping artifact (fixture · SHA-256 · pages)", "Digests still match disk",
         "What changed since the design was settled", "Owner note"],
        [22, 46, 7, 8, 12, 24, 44, 54, 12, 60, 30],
    )
    for r in pkg["batchAdoptionList"]:
        design = "; ".join(
            f"{d.get('role')}: {d.get('path')}" for d in (r.get("controllingLegalDesignRecord") or [])
            if d.get("role") in {"state_legal_design_memo", "legal_design_track_registry"}
        ) or "; ".join(d.get("path", "") for d in (r.get("controllingLegalDesignRecord") or [])[:2])
        art = r.get("currentShippingArtifact") or {}
        fixtures = "\n".join(
            f"{x.get('fixture')} · {(x.get('sha256OnDiskNow') or '')[:16]}… · {x.get('pageCount')}pp"
            for x in (art.get("fixtures") or [])
        ) or "not recorded"
        ws.append([
            "", r["familyId"], r.get("jurisdiction"), r.get("routeCount"),
            r.get("runtimeRepresentedRoutes"), r.get("deliveryType"), design, fixtures,
            "yes" if art.get("everyDeclaredDigestStillMatchesDisk") else "NO",
            r.get("exactDelta") or "", "",
        ])
    for row in ws.iter_rows(min_row=2):
        for c in row:
            c.alignment = WRAP if c.column in (7, 8, 10, 11) else TOP

    # ---- what the adoption asserts ----------------------------------------
    ws2 = sheet(wb, "What adoption asserts", ["", ""], [46, 110])
    ws2.append(["Status", pkg["status"]])
    ws2.append(["What is being asked", pkg["whatIsBeingAsked"]])
    ws2.append(["Applies nothing", pkg["appliesNothing"]])
    ws2.append(["Existing approval", pkg["theExistingApproval"]["approvalId"]])
    ws2.append(["Decision owner", pkg["theExistingApproval"]["decisionOwner"]])
    ws2.append(["Signature required", "no" if not pkg["theExistingApproval"]["requiresSignature"] else "yes"])
    ws2.append(["Why it does not already cover these", pkg["theExistingApproval"]["whyItDoesNotAlreadyCoverThese"]])
    ws2.append(["Unchanged on every family in the list",
                "remedy · eligibility · venue · filing destination · service · official-form strategy · substantive legal language"])
    for line in pkg["explicitNonGrants"]:
        ws2.append(["Explicit non-grant", line])
    for row in ws2.iter_rows(min_row=2):
        row[0].font = Font(bold=True, size=10)
        for c in row:
            c.alignment = WRAP

    # ---- exceptions --------------------------------------------------------
    ws3 = sheet(
        wb, "Excluded (25 families)",
        ["Why excluded", "Family", "State", "Routes", "Exact delta", "Question for the reviewer"],
        [30, 46, 7, 40, 60, 70],
    )
    exc = pkg["exceptionsRemovedFromTheBatch"]
    for r in exc["substantiveCounselReviewRequired"]:
        ws3.append(["Substantive counsel review", r["familyId"], r.get("jurisdiction"),
                    "\n".join(r.get("routeKeys") or []), r.get("exactDelta") or "",
                    r.get("substantiveQuestionForReviewer") or ""])
    for r in exc["unresolvedProductTreatment"]:
        ws3.append(["Unresolved product treatment", r["familyId"], r.get("jurisdiction"),
                    "\n".join(r.get("routeKeys") or []), r.get("exactDelta") or "",
                    r.get("productTreatmentQuestion") or ""])
    for row in ws3.iter_rows(min_row=2):
        for c in row:
            c.alignment = WRAP

    # ---- the eight questions ----------------------------------------------
    ws4 = sheet(wb, "The 8 open questions",
                ["#", "Question ID", "Bucket", "Question", "Families it covers", "Owner decision"],
                [5, 44, 8, 90, 50, 34])
    for i, q in enumerate(pkg["theEightQuestionsRequiringSeparateResolution"], start=1):
        if isinstance(q, dict):
            qid = str(q.get("id") or "")
            bucket = str(q.get("bucket") or "")
            text = str(q.get("question") or q.get("text") or "")
            fams = q.get("families") or q.get("familyIds") or []
        else:
            qid, bucket, text, fams = "", "", str(q), []
        if not isinstance(fams, list):
            fams = [str(fams)]
        ws4.append([i, qid, bucket, text, "\n".join(str(f) for f in fams), ""])
    for row in ws4.iter_rows(min_row=2):
        for c in row:
            c.alignment = WRAP

    out = Path(args.out)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    print(f"wrote {out}")
    print(f"  adopt sheet: {len(pkg['batchAdoptionList'])} famil(ies)")
    print(f"  excluded:    {len(exc['substantiveCounselReviewRequired'])} substantive + "
          f"{len(exc['unresolvedProductTreatment'])} unresolved")
    print(f"  questions:   {len(pkg['theEightQuestionsRequiringSeparateResolution'])}")


if __name__ == "__main__":
    main()
